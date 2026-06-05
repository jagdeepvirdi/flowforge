"""Tests for flowforge/steps/data_load.py — DataLoadStep run() method and helper methods.

All DB calls are mocked — no live database required.
"""
import csv
from unittest.mock import MagicMock, patch

import pytest

from flowforge.steps.data_load import DataLoadStep

# ── helpers ───────────────────────────────────────────────────────────────────

def make_step(config: dict) -> DataLoadStep:
    return DataLoadStep(name='test_load', config=config)


def _mock_conn(db_type='postgresql'):
    conn = MagicMock()
    conn.db_type = db_type
    conn.__enter__ = lambda s: s
    conn.__exit__ = MagicMock(return_value=False)
    conn.execute_query.return_value = []
    conn.make_placeholders.return_value = '%s, %s'
    conn.execute_many.return_value = 2
    return conn


# ── validation tests ──────────────────────────────────────────────────────────

def test_missing_target_connection_id():
    step = make_step({
        'target_table': 'my_table',
        'mode': 'append',
        'source': {'type': 'file', 'file_path': '/tmp/x.csv'},
    })
    result = step.run({})
    assert result.success is False
    assert 'required' in result.error.lower() or 'target_connection_id' in result.error


def test_missing_target_table():
    step = make_step({
        'target_connection_id': 'fake-uuid',
        'mode': 'append',
        'source': {'type': 'file', 'file_path': '/tmp/x.csv'},
    })
    result = step.run({})
    assert result.success is False
    assert 'target_table' in result.error.lower() or 'required' in result.error.lower()


def test_invalid_mode():
    step = make_step({
        'target_connection_id': 'fake-uuid',
        'target_table': 'my_table',
        'mode': 'upsert',  # invalid
        'source': {'type': 'file', 'file_path': '/tmp/x.csv'},
    })
    result = step.run({})
    assert result.success is False
    assert 'mode' in result.error.lower()


def test_invalid_source_type():
    step = make_step({
        'target_connection_id': 'fake-uuid',
        'target_table': 'my_table',
        'mode': 'append',
        'source': {'type': 'kafka'},  # invalid
    })
    result = step.run({})
    assert result.success is False
    assert 'source.type' in result.error.lower()


def test_source_read_exception():
    """When _load_source raises, returns StepResult(success=False, error='Source read failed...')."""
    step = make_step({
        'target_connection_id': 'fake-uuid',
        'target_table': 'my_table',
        'mode': 'append',
        'source': {'type': 'file', 'file_path': '/nonexistent/path.csv'},
    })
    result = step.run({})
    assert result.success is False
    assert 'source read failed' in result.error.lower()


def test_empty_rows_from_source():
    """0 rows from source → success=True, rows_affected=0."""
    step = make_step({
        'target_connection_id': 'fake-uuid',
        'target_table': 'my_table',
        'mode': 'append',
        'source': {'type': 'file', 'file_path': '/tmp/x.csv'},
    })
    with patch.object(step, '_load_source', return_value=(['id', 'name'], [])):
        result = step.run({})
    assert result.success is True
    assert result.rows_affected == 0


# ── column_map tests ──────────────────────────────────────────────────────────

def test_column_map_applied():
    """column_map renames source columns before insert."""
    step = make_step({
        'target_connection_id': 'fake-uuid',
        'target_table': 'my_table',
        'mode': 'append',
        'column_map': {'SRC_ID': 'id', 'SRC_NAME': 'name'},
        'source': {'type': 'file', 'file_path': '/tmp/x.csv'},
    })

    conn = _mock_conn()
    with patch.object(step, '_load_source', return_value=(['SRC_ID', 'SRC_NAME'], [(1, 'Alice')])), \
         patch('flowforge.connections.factory.get_connection', return_value=conn):
        result = step.run({})

    assert result.success is True
    # Verify the INSERT used mapped column names
    insert_call = conn.execute_many.call_args
    assert insert_call is not None
    insert_sql = insert_call[0][0]
    assert 'id' in insert_sql
    assert 'name' in insert_sql


def test_column_map_unmapped_columns_pass_through():
    """Columns not in column_map keep their original name."""
    step = make_step({
        'target_connection_id': 'fake-uuid',
        'target_table': 'my_table',
        'mode': 'append',
        'column_map': {'SRC_ID': 'id'},
        'source': {'type': 'file', 'file_path': '/tmp/x.csv'},
    })

    conn = _mock_conn()
    with patch.object(step, '_load_source', return_value=(['SRC_ID', 'value'], [(1, 42)])), \
         patch('flowforge.connections.factory.get_connection', return_value=conn):
        result = step.run({})

    assert result.success is True
    insert_sql = conn.execute_many.call_args[0][0]
    assert 'id' in insert_sql
    assert 'value' in insert_sql


# ── insert exception ──────────────────────────────────────────────────────────

def test_insert_exception_returns_failure():
    """Exception during DB insert → StepResult(success=False)."""
    step = make_step({
        'target_connection_id': 'fake-uuid',
        'target_table': 'my_table',
        'mode': 'append',
        'source': {'type': 'file', 'file_path': '/tmp/x.csv'},
    })

    conn = _mock_conn()
    conn.execute_many.side_effect = RuntimeError('insert failed')

    with patch.object(step, '_load_source', return_value=(['id'], [(1,)])), \
         patch('flowforge.connections.factory.get_connection', return_value=conn):
        result = step.run({})

    assert result.success is False
    assert 'insert failed' in result.error


# ── _load_file ────────────────────────────────────────────────────────────────

def test_load_file_nonexistent_raises_file_not_found():
    step = make_step({})
    from flowforge.engine.context import render as render_fn
    with pytest.raises(FileNotFoundError, match='not found'):
        step._load_file(
            {'type': 'file', 'file_path': '/nonexistent/path/data.csv'},
            {},
            render_fn,
        )


def test_load_file_unknown_format_raises_value_error(tmp_path):
    step = make_step({})
    from flowforge.engine.context import render as render_fn
    fake_file = tmp_path / 'data.jsonl'
    fake_file.write_text('{"id": 1}')
    with pytest.raises(ValueError, match='Unknown file format'):
        step._load_file(
            {'type': 'file', 'file_path': str(fake_file)},
            {},
            render_fn,
        )


def test_load_file_explicit_unknown_format_raises(tmp_path):
    step = make_step({})
    from flowforge.engine.context import render as render_fn
    fake_file = tmp_path / 'data.csv'
    fake_file.write_text('id,name\n1,Alice\n')
    with pytest.raises(ValueError, match='Unknown file format'):
        step._load_file(
            {'type': 'file', 'file_path': str(fake_file), 'file_format': 'parquet'},
            {},
            render_fn,
        )


# ── _read_csv ─────────────────────────────────────────────────────────────────

def test_read_csv_basic(tmp_path):
    f = tmp_path / 'data.csv'
    f.write_text('id,name,amount\n1,Alice,9.99\n2,Bob,19.99\n')
    step = make_step({})
    headers, rows = step._read_csv(f)
    assert headers == ['id', 'name', 'amount']
    assert len(rows) == 2
    assert rows[0] == ('1', 'Alice', '9.99')
    assert rows[1] == ('2', 'Bob', '19.99')


def test_read_csv_with_bom(tmp_path):
    """CSV with UTF-8 BOM is handled correctly."""
    f = tmp_path / 'bom.csv'
    # Write BOM + CSV
    content = '﻿id,name\n1,Alice\n'
    f.write_text(content, encoding='utf-8')
    step = make_step({})
    headers, rows = step._read_csv(f)
    assert headers[0] == 'id'  # BOM stripped


def test_read_csv_skips_empty_rows(tmp_path):
    f = tmp_path / 'sparse.csv'
    f.write_text('id,name\n1,Alice\n\n2,Bob\n')
    step = make_step({})
    _, rows = step._read_csv(f)
    assert len(rows) == 2  # empty row skipped


def test_read_csv_strips_header_whitespace(tmp_path):
    f = tmp_path / 'ws.csv'
    f.write_text(' id , name , amount \n1,Alice,10\n')
    step = make_step({})
    headers, _ = step._read_csv(f)
    assert headers == ['id', 'name', 'amount']


# ── _read_excel ───────────────────────────────────────────────────────────────

def test_read_excel_basic(tmp_path):
    """_read_excel reads headers and rows from xlsx."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip('openpyxl not installed')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['id', 'name', 'score'])
    ws.append([1, 'Alice', 9.5])
    ws.append([2, 'Bob', 8.0])
    path = tmp_path / 'data.xlsx'
    wb.save(str(path))
    wb.close()

    step = make_step({})
    headers, rows = step._read_excel(path, None)
    assert headers == ['id', 'name', 'score']
    assert len(rows) == 2


def test_read_excel_with_sheet_name(tmp_path):
    try:
        import openpyxl
    except ImportError:
        pytest.skip('openpyxl not installed')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'MySheet'
    ws.append(['col1', 'col2'])
    ws.append(['a', 'b'])
    path = tmp_path / 'named.xlsx'
    wb.save(str(path))
    wb.close()

    step = make_step({})
    headers, rows = step._read_excel(path, 'MySheet')
    assert headers == ['col1', 'col2']
    assert rows[0] == ('a', 'b')


def test_read_excel_skips_all_none_rows(tmp_path):
    try:
        import openpyxl
    except ImportError:
        pytest.skip('openpyxl not installed')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['id', 'val'])
    ws.append([1, 'x'])
    ws.append([None, None])  # all-null row should be skipped
    ws.append([2, 'y'])
    path = tmp_path / 'nulls.xlsx'
    wb.save(str(path))
    wb.close()

    step = make_step({})
    _, rows = step._read_excel(path, None)
    assert len(rows) == 2  # null row skipped


# ── _load_query ───────────────────────────────────────────────────────────────

def test_load_query_missing_connection_id_raises():
    step = make_step({})
    from flowforge.engine.context import render as render_fn
    with pytest.raises(ValueError, match='connection_id'):
        step._load_query({'type': 'query', 'query': 'SELECT 1'}, {}, render_fn)


def test_load_query_missing_query_raises():
    step = make_step({})
    from flowforge.engine.context import render as render_fn
    with pytest.raises(ValueError, match='query'):
        step._load_query({'type': 'query', 'connection_id': 'fake-id', 'query': ''}, {}, render_fn)


def test_load_query_success():
    step = make_step({})
    from flowforge.engine.context import render as render_fn

    mock_conn = MagicMock()
    mock_conn.__enter__ = lambda s: s
    mock_conn.__exit__ = MagicMock(return_value=False)
    mock_conn.execute_query_with_columns.return_value = (
        [(1, 'Alice'), (2, 'Bob')],
        ['id', 'name'],
    )

    with patch('flowforge.connections.factory.get_connection', return_value=mock_conn):
        columns, rows = step._load_query(
            {'type': 'query', 'connection_id': 'fake-id', 'query': 'SELECT id, name FROM t'},
            {},
            render_fn,
        )

    assert columns == ['id', 'name']
    assert len(rows) == 2


def test_load_query_renders_jinja_in_sql():
    step = make_step({})
    from flowforge.engine.context import render as render_fn

    mock_conn = MagicMock()
    mock_conn.__enter__ = lambda s: s
    mock_conn.__exit__ = MagicMock(return_value=False)
    mock_conn.execute_query_with_columns.return_value = ([], ['id'])

    captured_sql = {}

    def capture_query(sql):
        captured_sql['sql'] = sql
        return [], ['id']

    mock_conn.execute_query_with_columns.side_effect = capture_query

    ctx = {'current_month': '2026-05'}
    with patch('flowforge.connections.factory.get_connection', return_value=mock_conn):
        step._load_query(
            {
                'type': 'query',
                'connection_id': 'fake-id',
                'query': "SELECT * FROM t WHERE month = '{{ current_month }}'",
            },
            ctx,
            render_fn,
        )

    assert '2026-05' in captured_sql.get('sql', '')


# ── full run end-to-end ───────────────────────────────────────────────────────

def test_full_run_append_mode():
    """Full run in append mode inserts rows without TRUNCATE."""
    step = make_step({
        'target_connection_id': 'fake-uuid',
        'target_table': 'staging.sales',
        'mode': 'append',
        'source': {'type': 'file', 'file_path': '/fake/path.csv'},
    })
    conn = _mock_conn()

    with patch.object(step, '_load_source', return_value=(['id', 'amount'], [(1, 9.99), (2, 19.99)])), \
         patch('flowforge.connections.factory.get_connection', return_value=conn):
        result = step.run({})

    assert result.success is True
    assert result.rows_affected > 0
    # No TRUNCATE in append mode
    write_calls = [c[0][0] for c in conn.execute_write.call_args_list]
    assert not any('TRUNCATE' in c for c in write_calls)


def test_full_run_replace_mode():
    """Full run in replace mode calls TRUNCATE before insert."""
    step = make_step({
        'target_connection_id': 'fake-uuid',
        'target_table': 'staging.sales',
        'mode': 'replace',
        'source': {'type': 'file', 'file_path': '/fake/path.csv'},
    })
    conn = _mock_conn()

    with patch.object(step, '_load_source', return_value=(['id', 'amount'], [(1, 9.99)])), \
         patch('flowforge.connections.factory.get_connection', return_value=conn):
        result = step.run({})

    assert result.success is True
    write_calls = [c[0][0] for c in conn.execute_write.call_args_list]
    assert any('TRUNCATE' in c for c in write_calls)


def test_full_run_success_log_contains_table_name():
    step = make_step({
        'target_connection_id': 'fake-uuid',
        'target_table': 'reporting.monthly_revenue',
        'mode': 'append',
        'source': {'type': 'file', 'file_path': '/fake/path.csv'},
    })
    conn = _mock_conn()

    with patch.object(step, '_load_source', return_value=(['id'], [(1,), (2,)])), \
         patch('flowforge.connections.factory.get_connection', return_value=conn):
        result = step.run({})

    assert result.success is True
    assert 'reporting.monthly_revenue' in result.logs


def test_full_run_with_csv_file(tmp_path):
    """Full run reading from a real CSV file."""
    csv_file = tmp_path / 'sales.csv'
    csv_file.write_text('id,product,amount\n1,Widget,9.99\n2,Gadget,19.99\n3,Doohickey,4.49\n')

    step = make_step({
        'target_connection_id': 'fake-uuid',
        'target_table': 'staging.sales',
        'mode': 'append',
        'source': {'type': 'file', 'file_path': str(csv_file)},
    })
    conn = _mock_conn()
    conn.execute_many.return_value = 3

    with patch('flowforge.connections.factory.get_connection', return_value=conn):
        result = step.run({'steps': {}})

    assert result.success is True
    assert result.rows_affected == 3


def test_full_run_chunk_size_respected(tmp_path):
    """Large dataset with small chunk_size → execute_many called multiple times."""
    csv_file = tmp_path / 'big.csv'
    with open(csv_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['id', 'val'])
        for i in range(25):
            writer.writerow([i, f'val_{i}'])

    step = make_step({
        'target_connection_id': 'fake-uuid',
        'target_table': 'my_table',
        'mode': 'append',
        'chunk_size': 10,
        'source': {'type': 'file', 'file_path': str(csv_file)},
    })
    conn = _mock_conn()
    conn.execute_many.return_value = 10

    with patch('flowforge.connections.factory.get_connection', return_value=conn):
        result = step.run({'steps': {}})

    assert result.success is True
    # 25 rows / 10 per chunk = 3 calls
    assert conn.execute_many.call_count == 3
