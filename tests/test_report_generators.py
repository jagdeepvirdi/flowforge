"""Tests for Excel, CSV, PDF, and JSON report file generation."""
import csv
from pathlib import Path

import pytest

COLUMNS = ['id', 'name', 'amount', 'date']
ROWS = [
    (1, 'Alice', 1200.50, '2026-01-15'),
    (2, 'Bob',   980.00,  '2026-01-16'),
    (3, 'Carol', 2450.75, '2026-01-17'),
]


# ── CSV ───────────────────────────────────────────────────────────────────────

def test_csv_creates_file(tmp_path):
    from flowforge.reports.csv_report import generate
    out = tmp_path / 'report.csv'
    generate(ROWS, COLUMNS, out)
    assert out.exists()
    assert out.stat().st_size > 0


def test_csv_header_row(tmp_path):
    from flowforge.reports.csv_report import generate
    out = tmp_path / 'report.csv'
    generate(ROWS, COLUMNS, out)
    with open(out, encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        header = next(reader)
    assert header == COLUMNS


def test_csv_data_rows(tmp_path):
    from flowforge.reports.csv_report import generate
    out = tmp_path / 'report.csv'
    generate(ROWS, COLUMNS, out)
    with open(out, encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        next(reader)   # skip header
        data = list(reader)
    assert len(data) == 3
    assert data[0][1] == 'Alice'
    assert data[1][1] == 'Bob'


def test_csv_without_header(tmp_path):
    from flowforge.reports.csv_report import generate
    out = tmp_path / 'report.csv'
    generate(ROWS, COLUMNS, out, include_header=False)
    with open(out, encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        rows = list(reader)
    assert len(rows) == 3
    assert rows[0][1] == 'Alice'


def test_csv_custom_delimiter(tmp_path):
    from flowforge.reports.csv_report import generate
    out = tmp_path / 'report.tsv'
    generate(ROWS, COLUMNS, out, delimiter='\t')
    content = out.read_text(encoding='utf-8-sig')
    assert '\t' in content


def test_csv_empty_rows(tmp_path):
    from flowforge.reports.csv_report import generate
    out = tmp_path / 'empty.csv'
    generate([], COLUMNS, out)
    with open(out, encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        header = next(reader)
        data = list(reader)
    assert header == COLUMNS
    assert data == []


def test_csv_creates_parent_dirs(tmp_path):
    from flowforge.reports.csv_report import generate
    out = tmp_path / 'nested' / 'deep' / 'report.csv'
    generate(ROWS, COLUMNS, out)
    assert out.exists()


def test_csv_returns_output_path(tmp_path):
    from flowforge.reports.csv_report import generate
    out = tmp_path / 'report.csv'
    returned = generate(ROWS, COLUMNS, out)
    assert returned == out


# ── Excel ─────────────────────────────────────────────────────────────────────

pytest.importorskip('openpyxl', reason='openpyxl not installed')


def test_excel_creates_file(tmp_path):
    from flowforge.reports.excel_report import generate
    out = tmp_path / 'report.xlsx'
    generate(ROWS, COLUMNS, out)
    assert out.exists()
    assert out.stat().st_size > 0


def test_excel_header_row(tmp_path):
    from openpyxl import load_workbook

    from flowforge.reports.excel_report import generate
    out = tmp_path / 'report.xlsx'
    generate(ROWS, COLUMNS, out)
    wb = load_workbook(out)
    ws = wb.active
    headers = [ws.cell(1, i + 1).value for i in range(len(COLUMNS))]
    assert headers == COLUMNS


def test_excel_header_is_bold(tmp_path):
    from openpyxl import load_workbook

    from flowforge.reports.excel_report import generate
    out = tmp_path / 'report.xlsx'
    generate(ROWS, COLUMNS, out)
    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(1, 1).font.bold is True


def test_excel_data_rows(tmp_path):
    from openpyxl import load_workbook

    from flowforge.reports.excel_report import generate
    out = tmp_path / 'report.xlsx'
    generate(ROWS, COLUMNS, out)
    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(2, 2).value == 'Alice'
    assert ws.cell(3, 2).value == 'Bob'
    assert ws.cell(4, 2).value == 'Carol'


def test_excel_correct_row_count(tmp_path):
    from openpyxl import load_workbook

    from flowforge.reports.excel_report import generate
    out = tmp_path / 'report.xlsx'
    generate(ROWS, COLUMNS, out)
    wb = load_workbook(out)
    ws = wb.active
    assert ws.max_row == len(ROWS) + 1   # header + data rows


def test_excel_custom_sheet_name(tmp_path):
    from openpyxl import load_workbook

    from flowforge.reports.excel_report import generate
    out = tmp_path / 'report.xlsx'
    generate(ROWS, COLUMNS, out, sheet_name='Revenue')
    wb = load_workbook(out)
    assert 'Revenue' in wb.sheetnames


def test_excel_empty_rows(tmp_path):
    from openpyxl import load_workbook

    from flowforge.reports.excel_report import generate
    out = tmp_path / 'empty.xlsx'
    generate([], COLUMNS, out)
    wb = load_workbook(out)
    ws = wb.active
    assert ws.max_row == 1   # header only


def test_excel_creates_parent_dirs(tmp_path):
    from flowforge.reports.excel_report import generate
    out = tmp_path / 'nested' / 'report.xlsx'
    generate(ROWS, COLUMNS, out)
    assert out.exists()


def test_excel_returns_output_path(tmp_path):
    from flowforge.reports.excel_report import generate
    out = tmp_path / 'report.xlsx'
    returned = generate(ROWS, COLUMNS, out)
    assert returned == out


# ── JSON ──────────────────────────────────────────────────────────────────────

def test_json_creates_file(tmp_path):
    from flowforge.reports.json_report import generate
    out = tmp_path / 'report.json'
    generate(ROWS, COLUMNS, out)
    assert out.exists()
    assert out.stat().st_size > 0


def test_json_returns_output_path(tmp_path):
    from flowforge.reports.json_report import generate
    out = tmp_path / 'report.json'
    returned = generate(ROWS, COLUMNS, out)
    assert returned == out


def test_json_is_valid_json(tmp_path):
    import json

    from flowforge.reports.json_report import generate
    out = tmp_path / 'report.json'
    generate(ROWS, COLUMNS, out)
    data = json.loads(out.read_text(encoding='utf-8'))
    assert isinstance(data, list)
    assert len(data) == len(ROWS)


def test_json_keys_match_columns(tmp_path):
    import json

    from flowforge.reports.json_report import generate
    out = tmp_path / 'report.json'
    generate(ROWS, COLUMNS, out)
    data = json.loads(out.read_text(encoding='utf-8'))
    assert list(data[0].keys()) == COLUMNS


def test_json_values_correct(tmp_path):
    import json

    from flowforge.reports.json_report import generate
    out = tmp_path / 'report.json'
    generate(ROWS, COLUMNS, out)
    data = json.loads(out.read_text(encoding='utf-8'))
    assert data[0]['name'] == 'Alice'
    assert data[1]['id'] == 2


def test_json_empty_rows(tmp_path):
    import json

    from flowforge.reports.json_report import generate
    out = tmp_path / 'empty.json'
    generate([], COLUMNS, out)
    data = json.loads(out.read_text(encoding='utf-8'))
    assert data == []


def test_json_creates_parent_dirs(tmp_path):
    from flowforge.reports.json_report import generate
    out = tmp_path / 'nested' / 'deep' / 'report.json'
    generate(ROWS, COLUMNS, out)
    assert out.exists()


def test_json_custom_indent(tmp_path):
    from flowforge.reports.json_report import generate
    out = tmp_path / 'compact.json'
    generate(ROWS, COLUMNS, out, indent=0)
    # indent=0 → newlines but no spaces; file should be smaller than default
    content = out.read_text(encoding='utf-8')
    assert content.strip().startswith('[')


def test_json_non_serialisable_value(tmp_path):
    """default=str fallback handles non-JSON-serialisable types (dates, Decimal)."""
    import json
    from datetime import date
    from decimal import Decimal

    from flowforge.reports.json_report import generate
    rows = [(1, Decimal('9.99'), date(2026, 1, 1))]
    cols = ['id', 'price', 'date']
    out = tmp_path / 'special.json'
    generate(rows, cols, out)
    data = json.loads(out.read_text(encoding='utf-8'))
    assert data[0]['price'] == '9.99'
    assert data[0]['date'] == '2026-01-01'


# ── PDF (weasyprint mocked) ───────────────────────────────────────────────────

def test_pdf_generates_file(tmp_path):
    """PDF generation writes a file when WeasyPrint is mocked."""
    import sys
    from types import ModuleType
    from unittest.mock import MagicMock, patch

    mock_weasyprint = ModuleType('weasyprint')
    mock_html_instance = MagicMock()
    mock_html_instance.write_pdf = MagicMock(side_effect=lambda path: Path(path).write_bytes(b'%PDF'))
    mock_weasyprint.HTML = MagicMock(return_value=mock_html_instance)

    with patch.dict(sys.modules, {'weasyprint': mock_weasyprint}):
        from flowforge.reports.pdf_report import generate
        out = tmp_path / 'report.pdf'
        result = generate(ROWS, COLUMNS, out, title='Test Report')

    assert result == out
    assert out.exists()


def test_pdf_uses_default_template_when_no_template_path(tmp_path):
    import sys
    from types import ModuleType
    from unittest.mock import MagicMock, patch

    mock_weasyprint = ModuleType('weasyprint')
    captured = {}

    def capture_html(string=None, **kw):
        captured['html'] = string
        m = MagicMock()
        m.write_pdf = MagicMock(side_effect=lambda p: Path(p).write_bytes(b'%PDF'))
        return m

    mock_weasyprint.HTML = capture_html

    with patch.dict(sys.modules, {'weasyprint': mock_weasyprint}):
        from flowforge.reports.pdf_report import generate
        out = tmp_path / 'report.pdf'
        generate(ROWS, COLUMNS, out, title='My Title')

    assert 'My Title' in captured['html']
    assert 'Alice' in captured['html']


def test_pdf_uses_custom_template(tmp_path):
    import sys
    from types import ModuleType
    from unittest.mock import MagicMock, patch

    # Write a minimal custom template
    tmpl_file = tmp_path / 'custom.html'
    tmpl_file.write_text('<html><body><p>{{ title }}</p></body></html>', encoding='utf-8')

    mock_weasyprint = ModuleType('weasyprint')
    captured = {}

    def capture_html(string=None, **kw):
        captured['html'] = string
        m = MagicMock()
        m.write_pdf = MagicMock(side_effect=lambda p: Path(p).write_bytes(b'%PDF'))
        return m

    mock_weasyprint.HTML = capture_html

    with patch.dict(sys.modules, {'weasyprint': mock_weasyprint}):
        from flowforge.reports.pdf_report import generate
        out = tmp_path / 'out.pdf'
        generate(ROWS, COLUMNS, out, title='Custom', template_path=tmpl_file)

    assert '<p>Custom</p>' in captured['html']


def test_pdf_creates_parent_dirs(tmp_path):
    import sys
    from types import ModuleType
    from unittest.mock import MagicMock, patch

    mock_weasyprint = ModuleType('weasyprint')
    mock_html_instance = MagicMock()
    mock_html_instance.write_pdf = MagicMock(side_effect=lambda p: Path(p).write_bytes(b'%PDF'))
    mock_weasyprint.HTML = MagicMock(return_value=mock_html_instance)

    with patch.dict(sys.modules, {'weasyprint': mock_weasyprint}):
        from flowforge.reports.pdf_report import generate
        out = tmp_path / 'nested' / 'deep' / 'report.pdf'
        generate(ROWS, COLUMNS, out)

    assert out.exists()
