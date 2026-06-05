"""Excel report generator with optional per-column formatting rules.

column_formats is a list of dicts, one per column to format:
  {
    "column":        "Amount",          # matches header name (case-sensitive)
    "number_format": "#,##0.00",        # Excel format string
    "width":         18,                # column width in chars (overrides auto-fit)
    "conditional": [                    # optional list of conditional fill rules
      {
        "operator": "lt",               # lt | lte | gt | gte | eq | ne
        "value":    0,
        "bg_color": "FFC7CE",           # 6-char hex, no #
        "font_color": "9C0006"          # optional
      }
    ]
  }
"""
import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

_OPERATORS = {
    'lt':  lambda a, b: a <  b,
    'lte': lambda a, b: a <= b,
    'gt':  lambda a, b: a >  b,
    'gte': lambda a, b: a >= b,
    'eq':  lambda a, b: a == b,
    'ne':  lambda a, b: a != b,
}


def _matches_rule(value: Any, rule: dict) -> bool:
    op = _OPERATORS.get(rule.get('operator', ''))
    if op is None:
        return False
    try:
        return op(float(value), float(rule['value']))
    except (TypeError, ValueError):
        return False


def generate(
    rows: list[tuple],
    columns: list[str],
    output_path: Path,
    sheet_name: str = 'Sheet1',
    template_path: Path | None = None,
    column_formats: list[dict] | None = None,
) -> Path:
    """Write rows to an Excel file. Returns output_path."""
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Build lookup: column name → format rule
    fmt_by_col: dict[str, dict] = {}
    if column_formats:
        for rule in column_formats:
            col = rule.get('column', '')
            if col:
                fmt_by_col[col] = rule

    if template_path and template_path.exists():
        book = load_workbook(template_path)
        ws = book.active
        start_row = ws.max_row + 1
        # Derive column index mapping from existing header row
        col_idx_map: dict[str, int] = {}
        for cell in ws[1]:
            if cell.value and cell.value in fmt_by_col:
                col_idx_map[cell.value] = cell.column
    else:
        book = Workbook()
        ws = book.active
        ws.title = sheet_name
        header_fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
        col_idx_map = {}
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            col_idx_map[col_name] = col_idx
        start_row = 2

    # Write data rows
    for row_idx, row in enumerate(rows, start_row):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            col_name = columns[col_idx - 1] if col_idx <= len(columns) else None
            if col_name and col_name in fmt_by_col:
                rule = fmt_by_col[col_name]

                if rule.get('number_format'):
                    cell.number_format = rule['number_format']

                for cond in rule.get('conditional', []):
                    if _matches_rule(value, cond):
                        bg = cond.get('bg_color', '')
                        fc = cond.get('font_color', '')
                        if bg:
                            cell.fill = PatternFill(fill_type='solid', fgColor=bg)
                        if fc:
                            cell.font = Font(color=fc)
                        break  # first matching condition wins

    # Column widths: explicit override or auto-fit
    for col in ws.columns:
        col_letter = col[0].column_letter
        col_name   = columns[col[0].column - 1] if col[0].column <= len(columns) else None
        explicit_w = fmt_by_col.get(col_name or '', {}).get('width') if col_name else None
        if explicit_w:
            ws.column_dimensions[col_letter].width = int(explicit_w)
        else:
            max_len = max((len(str(cell.value or '')) for cell in col), default=0)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    book.save(output_path)
    book.close()
    logger.info("Excel report written: %s (%d rows)", output_path, len(rows))
    return output_path
