"""Multi-section report writer for health check steps (Excel and CSV)."""
import csv
import logging
from pathlib import Path

logger = logging.getLogger(__name__)

# Section = {'title': str, 'columns': list[str], 'rows': list[tuple]}


def write_health_report(sections: list[dict], output_path: Path, fmt: str = 'excel') -> Path:
    """Write a multi-section health report. Each section becomes one Excel sheet or CSV block."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if fmt == 'excel':
        return _write_excel(sections, output_path)
    if fmt == 'csv':
        return _write_csv(sections, output_path)
    raise ValueError(f"Unsupported health report format: {fmt!r}. Use 'excel' or 'csv'.")


def _write_excel(sections: list[dict], output_path: Path) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    header_fill = PatternFill(fill_type='solid', fgColor='2D3748')
    header_font = Font(bold=True, color='F1F5F9')

    for section in sections:
        if not section or not section.get('rows'):
            continue
        ws = wb.create_sheet(title=section['title'][:31])  # Excel sheet name cap

        for col_idx, col_name in enumerate(section['columns'], 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, row in enumerate(section['rows'], 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    if not wb.sheetnames:
        ws = wb.create_sheet('No Data')
        ws.cell(1, 1, 'No health metrics were collected.')

    wb.save(output_path)
    logger.info("Health report written: %s (%d sections)", output_path, len(sections))
    return output_path


def _write_csv(sections: list[dict], output_path: Path) -> Path:
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        first = True
        for section in sections:
            if not section or not section.get('rows'):
                continue
            if not first:
                writer.writerow([])
            writer.writerow([f'--- {section["title"]} ---'])
            writer.writerow(section['columns'])
            writer.writerows(section['rows'])
            first = False
    logger.info("Health report written: %s (%d sections)", output_path, len(sections))
    return output_path
